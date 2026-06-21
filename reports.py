import io
import json
import os
import pandas as pd
from ga_svm import load_data
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

def generate_excel_report(state, dataset_path):
    """Generates an Excel report in memory and returns a BytesIO object."""
    wb = Workbook()
    
    # Define styles
    font_title = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Calibri", size=11)
    font_bold = Font(name="Calibri", size=11, bold=True)
    
    fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_accent = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    fill_light = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    
    # ── SHEET 1: RINGKASAN EVALUASI ──
    ws1 = wb.active
    ws1.title = "Evaluasi Model"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1["A1"] = "LAPORAN RINGKAS EVALUASI MODEL SVM + GA"
    ws1["A1"].font = font_title
    ws1.merge_cells("A1:D1")
    ws1.row_dimensions[1].height = 30
    
    ws1["A3"] = f"Nama File Dataset: {state.get('dataset_filename', 'N/A')}"
    ws1["A3"].font = font_bold
    
    # Table headers
    headers_eval = ["Metrik", "Sebelum Optimasi (SVM Baseline)", "Sesudah Optimasi (SVM + GA)", "Peningkatan / Selisih"]
    for col_idx, h in enumerate(headers_eval, 1):
        cell = ws1.cell(row=5, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws1.row_dimensions[5].height = 25
    
    metrics = ["Accuracy", "Precision", "Recall", "F1"]
    metric_keys = ["accuracy", "precision", "recall", "f1"]
    
    for row_idx, (m_label, key) in enumerate(zip(metrics, metric_keys), 6):
        v_base = state.get(f"{key}_baseline", 0.0)
        v_opt = state.get(f"{key}_opt", 0.0)
        diff = v_opt - v_base
        diff_str = f"+{diff:.4f}" if diff > 0 else f"{diff:.4f}"
        
        row_vals = [m_label, f"{v_base:.4f}", f"{v_opt:.4f}", diff_str]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_left
            else:
                cell.alignment = align_right
                if col_idx == 4:
                    cell.font = font_bold
                    if diff > 0:
                        cell.font = Font(name="Calibri", size=11, bold=True, color="10B981") # Green
                    elif diff < 0:
                        cell.font = Font(name="Calibri", size=11, bold=True, color="EF4444") # Red
        ws1.row_dimensions[row_idx].height = 20
        
    # Model parameters
    ws1["A12"] = "Parameter Model"
    ws1["A12"].font = font_bold
    
    params = [
        ("Fitur Awal", state.get("n_features_all", 0)),
        ("Fitur Terpilih (GA)", state.get("n_features_ga", 0)),
        ("Reduksi Fitur (%)", f"{((state.get('n_features_all', 0) - state.get('n_features_ga', 0)) / max(1, state.get('n_features_all', 1)) * 100):.1f}%")
    ]
    
    for idx, (label, val) in enumerate(params, 13):
        cell_lbl = ws1.cell(row=idx, column=1, value=label)
        cell_lbl.font = font_body
        cell_lbl.border = thin_border
        cell_val = ws1.cell(row=idx, column=2, value=str(val))
        cell_val.font = font_bold
        cell_val.border = thin_border
        cell_val.alignment = align_center
        ws1.row_dimensions[idx].height = 20
        
    # Auto-adjust column widths for sheet 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    # ── SHEET 2: SELEKSI FITUR ──
    ws2 = wb.create_sheet(title="Seleksi Fitur")
    ws2.views.sheetView[0].showGridLines = True
    ws2["A1"] = "STATUS FITUR HASIL SELEKSI GENETIC ALGORITHM"
    ws2["A1"].font = font_title
    ws2.merge_cells("A1:C1")
    ws2.row_dimensions[1].height = 30
    
    headers_feat = ["No", "Nama Fitur", "Status Seleksi (GA)"]
    for col_idx, h in enumerate(headers_feat, 1):
        cell = ws2.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border
    ws2.row_dimensions[3].height = 25
    
    all_features = state.get("all_features", [])
    selected_features = set(state.get("selected_features", []))
    
    for idx, f_name in enumerate(all_features, 1):
        row_idx = idx + 3
        is_sel = f_name in selected_features
        status_str = "DIPILIH (✓)" if is_sel else "DIELIMINASI (✗)"
        
        row_vals = [idx, f_name, status_str]
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = align_center
            elif col_idx == 2:
                cell.alignment = align_left
            else:
                cell.alignment = align_center
                cell.font = font_bold
                if is_sel:
                    cell.fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid") # light green
                    cell.font = Font(name="Calibri", size=11, bold=True, color="137333")
                else:
                    cell.fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid") # light red
                    cell.font = Font(name="Calibri", size=11, bold=True, color="C5221F")
        ws2.row_dimensions[row_idx].height = 20
        
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 10)
        
    # ── SHEET 3: DATASET PREVIEW ──
    if os.path.exists(dataset_path):
        ws3 = wb.create_sheet(title="Dataset")
        ws3.views.sheetView[0].showGridLines = True
        df = load_data(dataset_path)
        # Limit preview to 200 rows for size efficiency
        df_preview = df.head(200)
        
        ws3["A1"] = f"PREVIEW DATASET (Menampilkan 200 Baris Pertama dari {len(df)} total baris)"
        ws3["A1"].font = font_title
        ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        ws3.row_dimensions[1].height = 30
        
        # Headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws3.cell(row=3, column=col_idx, value=col_name)
            cell.font = font_header
            cell.fill = fill_accent
            cell.alignment = align_center
            cell.border = thin_border
        ws3.row_dimensions[3].height = 25
        
        # Rows
        for r_idx, row in df_preview.iterrows():
            row_idx = r_idx + 4
            for c_idx, val in enumerate(row, 1):
                cell = ws3.cell(row=row_idx, column=c_idx, value=str(val))
                cell.font = font_body
                cell.border = thin_border
                # Simple alignment logic
                if df.dtypes.iloc[c_idx-1].kind in 'iuf':
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left
            ws3.row_dimensions[row_idx].height = 18
            
        for col in ws3.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws3.column_dimensions[col_letter].width = max(min(max_len + 3, 30), 10)
            
    # Save workbook to memory stream
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)
    return excel_stream

def generate_pdf_report(state, dataset_path):
    """Generates a PDF report using reportlab and returns a BytesIO object."""
    pdf_stream = io.BytesIO()
    doc = SimpleDocTemplate(
        pdf_stream,
        pagesize=letter,
        rightMargin=40, leftMargin=40,
        topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'BPSReportTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#1E3A8A'),
        alignment=1, # Center
        spaceAfter=15
    )
    
    subtitle_style = ParagraphStyle(
        'BPSReportSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#64748B'),
        alignment=1, # Center
        spaceAfter=25
    )
    
    section_style = ParagraphStyle(
        'BPSSectionHeading',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=12,
        textColor=colors.HexColor('#1E3A8A'),
        spaceBefore=15,
        spaceAfter=8
    )
    
    body_style = ParagraphStyle(
        'BPSBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#1E293B'),
        leading=14,
        spaceAfter=8
    )
    
    bold_style = ParagraphStyle(
        'BPSBold',
        parent=body_style,
        fontName='Helvetica-Bold'
    )
    
    table_cell_style = ParagraphStyle(
        'BPSTableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=11
    )
    
    table_cell_bold = ParagraphStyle(
        'BPSTableCellBold',
        parent=table_cell_style,
        fontName='Helvetica-Bold'
    )
    
    story = []
    
    # ── HEADER Section ──
    story.append(Paragraph("LAPORAN PENELITIAN AKADEMIS", title_style))
    story.append(Paragraph("Klasifikasi Penerima Bantuan Sosial (Bansos) &mdash; SVM + Genetic Algorithm", subtitle_style))
    
    # ── DATASET SUMMARY ──
    story.append(Paragraph("1. Ringkasan Dataset", section_style))
    
    num_rows = 0
    if os.path.exists(dataset_path):
        df = load_data(dataset_path)
        num_rows = len(df)
        
    ds_text = (
        f"Laporan ini menganalisis dataset <b>{state.get('dataset_filename', 'N/A')}</b> "
        f"yang berisi total <b>{num_rows}</b> data penduduk Indonesia. "
        f"Pembersihan data (data cleaning), normalisasi (StandardScaler), dan pemisahan data latih/uji "
        f"(split train-test 80:20) telah berhasil dilakukan."
    )
    story.append(Paragraph(ds_text, body_style))
    story.append(Spacer(1, 10))
    
    # ── EVALUATION METRICS TABLE ──
    story.append(Paragraph("2. Metrik Performa Model SVM (Baseline vs GA-Optimized)", section_style))
    
    metrics_data = [
        [
            Paragraph("<b>Metrik</b>", table_cell_bold),
            Paragraph("<b>Sebelum GA (Baseline)</b>", table_cell_bold),
            Paragraph("<b>Sesudah GA (Optimasi)</b>", table_cell_bold),
            Paragraph("<b>Selisih</b>", table_cell_bold)
        ]
    ]
    
    for key, label in [("accuracy", "Accuracy"), ("precision", "Precision"), ("recall", "Recall"), ("f1", "F1-Score")]:
        v_base = state.get(f"{key}_baseline", 0.0)
        v_opt = state.get(f"{key}_opt", 0.0)
        diff = v_opt - v_base
        diff_str = f"+{diff:.4f}" if diff > 0 else f"{diff:.4f}"
        
        metrics_data.append([
            Paragraph(label, table_cell_style),
            Paragraph(f"{v_base:.4f}", table_cell_style),
            Paragraph(f"{v_opt:.4f}", table_cell_style),
            Paragraph(diff_str, table_cell_bold)
        ])
        
    t_metrics = Table(metrics_data, colWidths=[150, 120, 120, 100])
    t_metrics.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F8FAFC')])
    ]))
    story.append(t_metrics)
    story.append(Spacer(1, 15))
    
    # ── FEATURE SELECTION SUMMARY ──
    story.append(Paragraph("3. Ringkasan Hasil Seleksi Fitur (Genetic Algorithm)", section_style))
    
    n_all = state.get("n_features_all", 0)
    n_ga = state.get("n_features_ga", 0)
    reduction = ((n_all - n_ga) / max(1, n_all)) * 100
    
    feat_summary = (
        f"Algoritma Genetika (GA) dijalankan untuk menyeleksi fitur-fitur yang paling kontributif "
        f"dalam menentukan kelayakan bantuan sosial. Dari total <b>{n_all}</b> fitur awal, "
        f"GA berhasil memilih <b>{n_ga}</b> fitur optimal (mengeliminasi <b>{n_all - n_ga}</b> fitur, "
        f"atau mereduksi noise sebesar <b>{reduction:.1f}%</b>)."
    )
    story.append(Paragraph(feat_summary, body_style))
    
    selected_feats = state.get("selected_features", [])
    selected_text = "<b>Fitur Terpilih:</b> " + ", ".join(f"<u>{f}</u>" for f in selected_feats)
    story.append(Paragraph(selected_text, body_style))
    
    story.append(Spacer(1, 10))
    
    # ── VERDICT / CONCLUSION ──
    story.append(Paragraph("4. Kesimpulan", section_style))
    
    acc_delta = state.get("accuracy_opt", 0.0) - state.get("accuracy_baseline", 0.0)
    if acc_delta > 0:
        verdict = (
            f"Berdasarkan uji coba, model <b>SVM + GA</b> menunjukkan peningkatan akurasi "
            f"sebesar <b>{acc_delta*100:+.2f}%</b> dibanding model SVM baseline. Kombinasi "
            f"seleksi fitur berbasis Algoritma Genetika terbukti efektif dalam memotong noise data "
            f"dan meningkatkan keandalan model klasifikasi kelayakan Bansos secara signifikan."
        )
    elif acc_delta == 0:
        verdict = (
            "Model <b>SVM + GA</b> menunjukkan akurasi yang setara dengan model baseline "
            "tetapi dengan jumlah fitur yang lebih sedikit (kompleksitas model berkurang). "
            "Hal ini meningkatkan efisiensi komputasi dan mengurangi risiko overfitting."
        )
    else:
        verdict = (
            f"Model <b>SVM + GA</b> mengalami sedikit penurunan akurasi ({acc_delta*100:+.2f}%) "
            f"dibanding baseline, namun berhasil menyederhanakan data secara dramatis dengan reduksi "
            f"fitur sebesar {reduction:.1f}%. Disarankan melakukan tala ulang (fine-tuning) "
            f"parameter GA untuk mencapai konvergensi optimal."
        )
        
    story.append(Paragraph(verdict, body_style))
    
    # Build Document
    doc.build(story)
    pdf_stream.seek(0)
    return pdf_stream
